from __future__ import annotations

import json
import math
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DATA_DIR = REPO_ROOT / "tools" / "battle_power" / "source" / "data"
RULES_PATH = REPO_ROOT / "tools" / "battle_power" / "selection_rules.json"
OUTPUT_PATH = REPO_ROOT / "app" / "src" / "main" / "assets" / "battle_power" / "character_coefficients.json"


@dataclass(frozen=True)
class CharacterRecord:
    uid: int
    name: str
    resource_name: str
    attack_type: str
    name_key: str
    active_skill_value_a: float
    ultimate_skill_value_a: float
    passive_value_a: float
    weight_value_a: float
    aside_value_a: float


def main() -> int:
    workbook_path = find_workbook_path()
    selection_rules = read_selection_rules()
    workbook = load_workbook(workbook_path, data_only=False)
    main_sheet = workbook.worksheets[0]
    character_sheet = workbook.worksheets[1]

    characters = read_character_records(character_sheet, selection_rules)
    default_character_uid = detect_default_character_uid(main_sheet, characters)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "defaultCharacterUid": default_character_uid,
        "characters": [
            {
                "uid": character.uid,
                "name": character.name,
                "resourceName": character.resource_name,
                "attackType": character.attack_type,
                "nameKey": character.name_key,
                "activeSkillValueA": format_number(character.active_skill_value_a),
                "ultimateSkillValueA": format_number(character.ultimate_skill_value_a),
                "passiveValueA": format_number(character.passive_value_a),
                "weightValueA": format_number(character.weight_value_a),
                "asideValueA": format_number(character.aside_value_a),
            }
            for character in characters
        ],
    }
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        "Synced battle power assets: "
        f"{len(characters)} characters, default uid {default_character_uid}.",
    )
    return 0


def find_workbook_path() -> Path:
    workbook_paths = [path for path in SOURCE_DATA_DIR.glob("*.xlsx") if not path.name.startswith("~$")]
    if len(workbook_paths) != 1:
        raise SystemExit(
            "Expected exactly one battle power workbook in source/data, "
            f"found {len(workbook_paths)}.",
        )
    return workbook_paths[0]


def read_selection_rules() -> dict:
    if not RULES_PATH.exists():
        raise SystemExit(f"Missing selection rules: {RULES_PATH}")
    return json.loads(RULES_PATH.read_text(encoding="utf-8"))


def read_character_records(sheet, selection_rules: dict) -> list[CharacterRecord]:
    max_uid_exclusive = int(selection_rules["max_uid_exclusive"])
    excluded_uids = {int(uid) for uid in selection_rules.get("excluded_uids", [])}
    characters: list[CharacterRecord] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        uid = normalize_int(row[0])
        if uid is None or uid >= max_uid_exclusive or uid in excluded_uids:
            continue

        name = normalize_text(row[1])
        resource_name = normalize_text(row[2])
        attack_type = normalize_text(row[3])
        name_key = normalize_text(row[4])
        if not name or not resource_name or not attack_type or not name_key:
            raise SystemExit(f"Incomplete character row at {sheet.title} row {row_index}.")

        characters.append(
            CharacterRecord(
                uid=uid,
                name=name,
                resource_name=resource_name,
                attack_type=attack_type,
                name_key=name_key,
                active_skill_value_a=normalize_float(row[5]),
                ultimate_skill_value_a=normalize_float(row[6]),
                passive_value_a=normalize_float(row[7]),
                weight_value_a=normalize_float(row[8]),
                aside_value_a=normalize_float(row[9]),
            ),
        )

    if not characters:
        raise SystemExit("No characters remained after applying selection rules.")

    ensure_unique(values=[character.uid for character in characters], label="character uid")
    ensure_unique(values=[character.name for character in characters], label="character name")
    return sorted(characters, key=lambda character: character.uid)


def detect_default_character_uid(main_sheet, characters: list[CharacterRecord]) -> int:
    attack_type = normalize_text(main_sheet.cell(row=2, column=3).value)
    active = normalize_float(main_sheet.cell(row=16, column=3).value)
    ultimate = normalize_float(main_sheet.cell(row=17, column=3).value)
    passive = normalize_float(main_sheet.cell(row=18, column=3).value)
    weight = normalize_float(main_sheet.cell(row=19, column=3).value)
    aside = normalize_float(main_sheet.cell(row=21, column=3).value)

    matches = [
        character
        for character in characters
        if character.attack_type == attack_type
        and nearly_equal(character.active_skill_value_a, active)
        and nearly_equal(character.ultimate_skill_value_a, ultimate)
        and nearly_equal(character.passive_value_a, passive)
        and nearly_equal(character.weight_value_a, weight)
        and nearly_equal(character.aside_value_a, aside)
    ]
    if len(matches) != 1:
        raise SystemExit(
            "Expected exactly one default character match from the main sheet, "
            f"found {len(matches)}.",
        )
    return matches[0].uid


def ensure_unique(values: list[object], label: str) -> None:
    if len(values) != len(set(values)):
        raise SystemExit(f"Duplicate {label} detected after filtering.")


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_int(value: object) -> int | None:
    if value is None:
        return None
    return int(value)


def normalize_float(value: object) -> float:
    if value is None:
        raise SystemExit("Unexpected empty numeric cell in battle power workbook.")
    return float(value)


def nearly_equal(left: float, right: float, tolerance: float = 1e-9) -> bool:
    return math.isclose(left, right, rel_tol=tolerance, abs_tol=tolerance)


def format_number(value: float) -> str:
    return format(value, ".15g")


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as error:  # pragma: no cover - defensive command-line fallback
        print(f"Failed to sync battle power assets: {error}", file=sys.stderr)
        raise SystemExit(1)
