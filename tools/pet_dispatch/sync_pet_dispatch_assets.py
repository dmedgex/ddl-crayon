from __future__ import annotations

import json
import shutil
import sys
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_ROOT = REPO_ROOT / "tools" / "pet_dispatch" / "source"
SOURCE_DATA_DIR = SOURCE_ROOT / "data"
SOURCE_IMAGES_DIR = SOURCE_ROOT / "images"
OUTPUT_ROOT = REPO_ROOT / "app" / "src" / "main" / "assets" / "pet_dispatch"
OUTPUT_IMAGES_DIR = OUTPUT_ROOT / "images"
OUTPUT_CATALOG_PATH = OUTPUT_ROOT / "catalog.json"

PET_WORKBOOK_NAME = "宠物列表.xlsx"
REGION_WORKBOOK_NAME = "跑腿地区.xlsx"
PET_SHEET_NAME = "Sheet1"
REGION_SHEET_NAME = "Sheet1"

RARITY_CODES = {
    "普通宠物": "COMMON",
    "高级宠物": "ADVANCED",
    "稀有宠物": "RARE",
    "传说宠物": "LEGENDARY",
}

SKILL_LEVEL_CODES = {"C", "B", "A", "S"}
SKILL_SCORE_MAP = {"C": 7, "B": 12, "A": 17, "S": 22}


@dataclass(frozen=True)
class PetRecord:
    id: int
    name: str
    rarity_code: str
    base_score: int
    skills: list[dict]
    image_asset_name: str


def main() -> int:
    pet_source_path = SOURCE_DATA_DIR / PET_WORKBOOK_NAME
    region_source_path = SOURCE_DATA_DIR / REGION_WORKBOOK_NAME

    validate_source_paths(
        paths=[
            pet_source_path,
            region_source_path,
            SOURCE_IMAGES_DIR,
        ],
    )

    pets = read_pets(pet_source_path)
    regions = read_regions(region_source_path)
    validate_image_mapping(pets)

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    rebuild_output_images_dir(pets)
    write_catalog(pets, regions)

    print(
        "Synced pet dispatch assets: "
        f"{len(pets)} pets, {len(regions)} regions, {sum(len(region['tasks']) for region in regions)} tasks.",
    )
    return 0


def validate_source_paths(paths: list[Path]) -> None:
    missing = [path for path in paths if not path.exists()]
    if missing:
        lines = "\n".join(f"- {path}" for path in missing)
        raise SystemExit(f"Missing required source paths:\n{lines}")


def read_pets(workbook_path: Path) -> list[PetRecord]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[PET_SHEET_NAME]

    pets: list[PetRecord] = []
    seen_names: set[str] = set()

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_name = normalize_text(row[0])
        if raw_name is None:
            continue
        if raw_name in seen_names:
            raise SystemExit(f"Duplicate pet name found in {workbook_path.name}: {raw_name}")
        seen_names.add(raw_name)

        raw_rarity = normalize_text(row[1])
        if raw_rarity not in RARITY_CODES:
            raise SystemExit(
                f"Unsupported rarity for pet '{raw_name}' in row {row_index}: {raw_rarity!r}",
            )

        skills = build_pet_skills(
            pet_name=raw_name,
            row_index=row_index,
            raw_skill_1=normalize_text(row[2]),
            raw_skill_1_level=normalize_text(row[3]),
            raw_skill_2=normalize_text(row[4]),
            raw_skill_2_level=normalize_text(row[5]),
        )

        pets.append(
            PetRecord(
                id=row_index - 1,
                name=raw_name,
                rarity_code=RARITY_CODES[raw_rarity],
                base_score=rarity_base_score(raw_rarity),
                skills=skills,
                image_asset_name=f"{raw_name}.png",
            ),
        )

    if not pets:
        raise SystemExit(f"No pets found in {workbook_path.name}.")
    return pets


def build_pet_skills(
    pet_name: str,
    row_index: int,
    raw_skill_1: str | None,
    raw_skill_1_level: str | None,
    raw_skill_2: str | None,
    raw_skill_2_level: str | None,
) -> list[dict]:
    skills: list[dict] = []
    for skill_name, level in (
        (raw_skill_1, raw_skill_1_level),
        (raw_skill_2, raw_skill_2_level),
    ):
        if skill_name is None and level is None:
            continue
        if not skill_name or not level:
            raise SystemExit(
                f"Incomplete skill definition for pet '{pet_name}' in row {row_index}.",
            )
        if level not in SKILL_LEVEL_CODES:
            raise SystemExit(
                f"Unsupported skill level for pet '{pet_name}' in row {row_index}: {level!r}",
            )
        skills.append(
            {
                "name": skill_name,
                "level": level,
                "score": SKILL_SCORE_MAP[level],
            },
        )
    return skills


def read_regions(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[REGION_SHEET_NAME]

    region_tasks: OrderedDict[str, list[dict]] = OrderedDict()
    current_region: str | None = None

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_region = normalize_text(row[0])
        if raw_region:
            current_region = raw_region
            region_tasks.setdefault(current_region, [])
        if current_region is None:
            raise SystemExit(
                f"Encountered task row before any region name in {workbook_path.name} row {row_index}.",
            )

        area = normalize_text(row[1])
        task = normalize_text(row[2])
        bonus_1 = normalize_text(row[3])
        bonus_2 = normalize_text(row[4])

        if area is None and task is None and bonus_1 is None and bonus_2 is None:
            continue
        if area is None or task is None:
            raise SystemExit(
                f"Incomplete task definition in region '{current_region}' at row {row_index}.",
            )

        region_tasks[current_region].append(
            {
                "id": len(region_tasks[current_region]),
                "area": area,
                "task": task,
                "bonusSkills": [value for value in (bonus_1, bonus_2) if value],
            },
        )

    if not region_tasks:
        raise SystemExit(f"No regions found in {workbook_path.name}.")

    return [
        {
            "name": region_name,
            "tasks": tasks,
        }
        for region_name, tasks in region_tasks.items()
    ]


def validate_image_mapping(pets: list[PetRecord]) -> None:
    expected_names = {pet.name for pet in pets}
    actual_names = {path.stem for path in SOURCE_IMAGES_DIR.glob("*.png")}

    missing_images = sorted(expected_names - actual_names)
    extra_images = sorted(actual_names - expected_names)

    if missing_images:
        raise SystemExit(
            "Missing pet images:\n" + "\n".join(f"- {name}.png" for name in missing_images),
        )
    if extra_images:
        raise SystemExit(
            "Found images without matching pet rows:\n" +
            "\n".join(f"- {name}.png" for name in extra_images),
        )


def rebuild_output_images_dir(pets: list[PetRecord]) -> None:
    if OUTPUT_IMAGES_DIR.exists():
        shutil.rmtree(OUTPUT_IMAGES_DIR)
    OUTPUT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    for pet in pets:
        source_image_path = SOURCE_IMAGES_DIR / pet.image_asset_name
        if not source_image_path.exists():
            raise SystemExit(f"Source image disappeared during sync: {source_image_path}")
        shutil.copy2(source_image_path, OUTPUT_IMAGES_DIR / pet.image_asset_name)


def write_catalog(pets: list[PetRecord], regions: list[dict]) -> None:
    payload = {
        "pets": [
            {
                "id": pet.id,
                "name": pet.name,
                "rarity": pet.rarity_code,
                "baseScore": pet.base_score,
                "skills": pet.skills,
                "imageAssetName": pet.image_asset_name,
            }
            for pet in pets
        ],
        "regions": regions,
    }
    OUTPUT_CATALOG_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def rarity_base_score(rarity_name: str) -> int:
    return {
        "普通宠物": 2,
        "高级宠物": 2,
        "稀有宠物": 3,
        "传说宠物": 5,
    }[rarity_name]


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as error:  # pragma: no cover - defensive command-line fallback
        print(f"Failed to sync pet dispatch assets: {error}", file=sys.stderr)
        raise SystemExit(1)
